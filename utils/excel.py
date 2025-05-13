import openpyxl
from openpyxl import Workbook
import os

class ExcelOperations:
    def __init__(self, file_path="soru_bankasi.xlsx"):
        self.file_path = file_path
        if not os.path.exists(file_path):  # Dosya yoksa oluştur
            workbook = Workbook()
            workbook.active.append(["ID", "Soru", "1. Seçenek", "2. Seçenek", "3. Seçenek", "4. Seçenek", "Doğru Cevap"])  # Başlık satırı ekle
            workbook.save(file_path)
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    def read_data(self):
        return list(self.sheet.iter_rows(values_only=True))[1:]  # Başlık satırını atla

    def add_data(self, row_data):
        self.sheet.append(row_data)
        self.workbook.save(self.file_path)

    def delete_row(self, row_index):
        self.sheet.delete_rows(row_index + 2)  # Başlık satırını hesaba kat
        self.workbook.save(self.file_path)

    def close(self):
        self.workbook.close()
